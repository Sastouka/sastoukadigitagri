# -*- coding: utf-8 -*-
import os
import sys
import json
import uuid
import hashlib
import platform
import datetime
import pandas as pd
import requests  # Pour l'API PayPal

from flask import (
    Flask, request, redirect, url_for, flash, send_file,
    render_template_string
)
from openpyxl import Workbook, load_workbook

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
import base64  # AJOUTÉ pour encoder les images en base64

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, LongTable, TableStyle, Table, Image, PageBreak
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER
from reportlab.lib.units import cm

app = Flask(__name__)
app.secret_key = "UNE_SUPER_CLE_SECRETE_FLASK"

# =============================================================================
# Partie Activation & Administration
# =============================================================================

# --- Paramètres d'activation ---
SECRET_SALT = "VOTRE_SEL_SECRET_UNIQUE"  # Remplacez par votre sel secret unique

def get_hardware_id():
    hardware_id = str(uuid.getnode())
    return hashlib.sha256(hardware_id.encode()).hexdigest()[:16]

def generate_activation_key_for_user(user_hardware_id, plan):
    return hashlib.sha256((user_hardware_id + SECRET_SALT + plan).encode()).hexdigest()[:15]

# =============================================================================
# Gestion de l'activation (plans : essai 7 jours, 1 an, illimité)
# =============================================================================

if platform.system() == "Windows":
    ACTIVATION_DIR = os.path.join(os.environ.get('APPDATA'), 'SystemData')
else:
    ACTIVATION_DIR = os.path.join(os.path.expanduser('~'), '.systemdata')
os.makedirs(ACTIVATION_DIR, exist_ok=True)
ACTIVATION_FILE = os.path.join(ACTIVATION_DIR, 'activation3264.json')

def check_activation():
    if not os.path.exists(ACTIVATION_FILE):
        activation_data = {
            "plan": "essai_7jours",
            "activation_date": datetime.date.today().isoformat()
        }
        with open(ACTIVATION_FILE, "w", encoding="utf-8") as f:
            json.dump(activation_data, f)
        return True
    else:
        with open(ACTIVATION_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        plan = data.get("plan")
        try:
            activation_date = datetime.date.fromisoformat(data.get("activation_date"))
        except Exception:
            return False
        if plan == "essai_7jours":
            return datetime.date.today() <= activation_date + datetime.timedelta(days=7)
        elif plan == "1 an":
            expected = generate_activation_key_for_user(get_hardware_id(), plan)
            if data.get("activation_code") == expected:
                try:
                    anniversary = activation_date.replace(year=activation_date.year + 1)
                except ValueError:
                    anniversary = activation_date + datetime.timedelta(days=365)
                return datetime.date.today() <= anniversary
            else:
                return False
        elif plan == "illimité":
            expected = generate_activation_key_for_user(get_hardware_id(), plan)
            return data.get("activation_code") == expected
        else:
            return False

def update_activation_after_payment(plan):
    data = {
        "plan": plan,
        "activation_date": datetime.date.today().isoformat(),
        "activation_code": generate_activation_key_for_user(get_hardware_id(), plan)
    }
    with open(ACTIVATION_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f)

# =============================================================================
# Nouvelle fonctionnalité : Contrôle de la période d'essai
# =============================================================================

def check_trial_period():
    if platform.system() == "Windows":
        hidden_folder = os.path.join(os.environ.get('APPDATA'), 'SystemData')
        os.makedirs(hidden_folder, exist_ok=True)
        file_path = os.path.join(hidden_folder, 'windows32')
    else:
        hidden_folder = os.path.join(os.path.expanduser('~'), '.systemdata')
        os.makedirs(hidden_folder, exist_ok=True)
        file_path = os.path.join(hidden_folder, 'windows32')
    from datetime import datetime
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding="utf-8") as f:
                stored_date_str = f.read().strip()
            stored_date = datetime.strptime(stored_date_str, '%Y-%m-%d')
        except Exception as e:
            flash("Le fichier de licence est corrompu. Veuillez contacter le support.", "error")
            return False
        if stored_date > datetime.now():
            flash("Le fichier de licence est corrompu ou la date système a été modifiée.", "error")
            return False
        days_passed = (datetime.now() - stored_date).days
        if days_passed > 150:
            flash("La période d'essai de 7 jours est terminée.<br>Contactez sastoukadigital@gmail.com ou Whatsapp au +212652084735.", "error")
            return False
        return True
    else:
        current_date_str = datetime.now().strftime('%Y-%m-%d')
        try:
            with open(file_path, 'w', encoding="utf-8") as f:
                f.write(current_date_str)
            if platform.system() == "Windows":
                import ctypes
                FILE_ATTRIBUTE_HIDDEN = 0x02
                FILE_ATTRIBUTE_SYSTEM = 0x04
                attrs = FILE_ATTRIBUTE_HIDDEN | FILE_ATTRIBUTE_SYSTEM
                ctypes.windll.kernel32.SetFileAttributesW(file_path, attrs)
            else:
                os.chmod(file_path, 0)
        except Exception as e:
            flash("Impossible de créer le fichier de licence.", "error")
            return False
        return True

@app.before_request
def enforce_trial_period():
    if request.endpoint not in ("activation", "activate", "purchase_plan", "paypal_success", "paypal_cancel", "change_theme", "trial_expired", "static"):
        if not check_trial_period():
            return redirect(url_for("trial_expired"))

@app.route("/trial_expired")
def trial_expired():
    return render_template_string("""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
      <meta charset="UTF-8"/>
      <title>Période d'essai expirée</title>
      <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="bg-light">
      <div class="container my-5">
        <div class="alert alert-danger" role="alert">
          La période d'essai de 7 jours est terminée.<br>
          Veuillez contacter <a href="mailto:sastoukadigital@gmail.com">sastoukadigital@gmail.com</a> ou Whatsapp au +212652084735.
        </div>
      </div>
    </body>
    </html>
    """)

# =============================================================================
# Partie Fichiers, Thèmes, Numéros de bons, PDF (inchangés)
# =============================================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
AHABIAFILES_DIR = os.path.join(BASE_DIR, "AHABIAFILES")
EXCEL_DIR = os.path.join(AHABIAFILES_DIR, "Excel")
PDF_LIVRAISON_DIR = os.path.join(AHABIAFILES_DIR, "PDF_Livraison")
PDF_STATS_DIR = os.path.join(AHABIAFILES_DIR, "PDF_Stats")
for d in [AHABIAFILES_DIR, EXCEL_DIR, PDF_LIVRAISON_DIR, PDF_STATS_DIR]:
    os.makedirs(d, exist_ok=True)

VOUCHER_FILE = os.path.join(AHABIAFILES_DIR, "last_voucher.txt")
USER_THEME_FILE = os.path.join(AHABIAFILES_DIR, "user_theme.json")

fruit_themes = {
    "Myrtille": {"bg": "#d0f0c0", "accent": "#4B0082", "fg": "#2f4f4f"},
    "Fraise": {"bg": "#ffe4e1", "accent": "#ff0000", "fg": "#800000"},
    "Framboise": {"bg": "#fce4ec", "accent": "#d81b60", "fg": "#880e4f"},
    "Avocat": {"bg": "#e8f5e9", "accent": "#388e3c", "fg": "#1b5e20"},
    "Pêche": {"bg": "#fff3e0", "accent": "#fb8c00", "fg": "#e65100"},
    "Nectarine": {"bg": "#f3e5f5", "accent": "#8e24aa", "fg": "#4a148c"},
    "Pomme": {"bg": "#e3f2fd", "accent": "#1976d2", "fg": "#0d47a1"}
}
DEFAULT_FRUIT = "Myrtille"

def load_user_theme():
    if os.path.exists(USER_THEME_FILE):
        try:
            with open(USER_THEME_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data.get("fruit", DEFAULT_FRUIT)
        except Exception:
            return DEFAULT_FRUIT
    return DEFAULT_FRUIT

def save_user_theme(fruit):
    try:
        with open(USER_THEME_FILE, "w", encoding="utf-8") as f:
            json.dump({"fruit": fruit}, f)
    except Exception:
        pass

def load_last_voucher_number():
    if os.path.exists(VOUCHER_FILE):
        with open(VOUCHER_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None

def save_last_voucher_number(num):
    with open(VOUCHER_FILE, "w", encoding="utf-8") as f:
        f.write(num)

def get_voucher_sequence(date_str):
    last_bon = load_last_voucher_number()
    if last_bon and last_bon.startswith("BL" + date_str):
        last_seq = int(last_bon[-2:])
        seq = last_seq + 1
    else:
        seq = 1
    return str(seq).zfill(2)

def generate_voucher_number(farmer):
    parts = farmer.split()
    if len(parts) >= 2:
        nom = parts[0][:2].upper().ljust(2, "X")
        prenom = parts[1][:2].upper().ljust(2, "X")
    else:
        nom = farmer[:2].upper().ljust(2, "X")
        prenom = farmer[:2].upper().ljust(2, "X")
    date_str = datetime.datetime.now().strftime("%d%m%Y")
    seq = get_voucher_sequence(date_str)
    num_bon = "BL" + date_str + nom + prenom + seq
    save_last_voucher_number(num_bon)
    return num_bon

def get_report_sequence(date_str):
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    seq = 1
    if os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
            if "HistoriqueRapports" in wb.sheetnames:
                ws = wb["HistoriqueRapports"]
                sequences = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and str(row[0]).startswith("R" + date_str):
                        try:
                            seq_num = int(str(row[0])[len("R" + date_str):])
                            sequences.append(seq_num)
                        except:
                            pass
                if sequences:
                    seq = max(sequences) + 1
        except Exception:
            seq = 1
    return str(seq).zfill(2)

def generate_report_number(farmer):
    parts = farmer.split()
    if len(parts) >= 2:
        nom = parts[0][:2].upper().ljust(2, "X")
        prenom = parts[1][:2].upper().ljust(2, "X")
    else:
        nom = farmer[:2].upper().ljust(2, "X")
        prenom = farmer[:2].upper().ljust(2, "X")
    date_str = datetime.datetime.now().strftime("%d%m%Y")
    seq = get_report_sequence(date_str)
    return "R" + date_str + nom + prenom + seq

class PDFGenerator:
    @staticmethod
    def generate_delivery_pdf(data, pdf_path):
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            leftMargin=1.5*cm,
            rightMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        styles = getSampleStyleSheet()
        styles["Normal"].fontSize = 14
        styles["Normal"].alignment = TA_JUSTIFY
        styles["Title"].fontSize = 14
        styles["Title"].alignment = TA_CENTER
        styles["Heading2"].fontSize = 14
        styles["Heading2"].alignment = TA_CENTER
        elements = []
        title = Paragraph("Gestion des Récoltes de " + data.get("fruit", "FRUIT") + " ENNAJIHI NAWFAL", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        invoice_paragraph = Paragraph(data['num_bon'], styles['Heading2'])
        elements.append(invoice_paragraph)
        elements.append(Spacer(1, 12))
        table_data = []
        table_data.append(["Champ", "Valeur"])
        for field, val in data['fields']:
            table_data.append([field, str(val)])
        PAGE_WIDTH, PAGE_HEIGHT = landscape(A4)
        usable_width = PAGE_WIDTH - (doc.leftMargin + doc.rightMargin)
        table = LongTable(
            table_data,
            colWidths=[0.4 * usable_width, 0.6 * usable_width],
            repeatRows=1
        )
        table.hAlign = 'CENTER'
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.beige]),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)
        doc.build(elements)

    @staticmethod
    def generate_stats_pdf(elements, pdf_path):
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=landscape(A4),
            leftMargin=1.5*cm,
            rightMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        doc.build(elements)

column_order = [
    "Date (JJ/MM/AAAA)",
    "Agriculteur",
    "Parcelle",
    "Produit",
    "Variété",
    "Nb Ouvriers Cueilleurs",
    "Nb Ouvriers Indirect",
    "Nb Ouvriers Autres",
    "Total Ouvriers",
    "Nombre Caporaux",
    "Poids Total Cueillis (kg)",
    "Écarts (Produit Déchet) en kg",
    "Poids Global"
]
column_abbr = {
    "Date (JJ/MM/AAAA)": "Date",
    "Agriculteur": "Agr.",
    "Parcelle": "Parc.",
    "Produit": "Prod.",
    "Variété": "Var.",
    "Nb Ouvriers Cueilleurs": "Ouv. C.",
    "Nb Ouvriers Indirect": "Ouv. I.",
    "Nb Ouvriers Autres": "Ouv. A.",
    "Total Ouvriers": "Tot. Ouv.",
    "Nombre Caporaux": "Caporaux",
    "Poids Total Cueillis (kg)": "Poids Total",
    "Écarts (Produit Déchet) en kg": "Écarts",
    "Poids Global": "P Global"
}
all_columns_extended = column_order.copy()
if "Mois" not in all_columns_extended:
    all_columns_extended.insert(1, "Mois")
if "Annee" not in all_columns_extended:
    all_columns_extended.insert(2, "Annee")

# =============================================================================
# Partie PayPal et Achat de Plans
# =============================================================================

PAYPAL_CLIENT_ID = "ATyh7nhaFjHLqrD4Bvp1Y2tXLeRub-9733ONYXASKr0sq6YEvbZm1QjcToKzFVRv6dIcGmyudbZT6YyL"
PAYPAL_SECRET = "EPysjDOTBgxhecho8xFualacKDeJn9udQebusanBYglTaBnW5lOT-Tg2v3gN5es_UJXXOGCVO0RG24bN"
PAYPAL_OAUTH_URL = "https://api-m.sandbox.paypal.com/v1/oauth2/token"
PAYPAL_ORDER_API = "https://api-m.sandbox.paypal.com/v2/checkout/orders"

def get_paypal_access_token():
    response = requests.post(
        PAYPAL_OAUTH_URL,
        headers={"Accept": "application/json", "Accept-Language": "en_US"},
        data={"grant_type": "client_credentials"},
        auth=(PAYPAL_CLIENT_ID, PAYPAL_SECRET)
    )
    if response.status_code == 200:
        return response.json()["access_token"]
    else:
        raise Exception(f"Erreur obtention token PayPal: {response.status_code} {response.text}")

def create_paypal_order(amount, currency="USD"):
    token = get_paypal_access_token()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    body = {
        "intent": "CAPTURE",
        "purchase_units": [
            {
                "amount": {
                    "currency_code": currency,
                    "value": amount
                }
            }
        ],
        "application_context": {
            "return_url": url_for("paypal_success", _external=True),
            "cancel_url": url_for("paypal_cancel", _external=True)
        }
    }
    response = requests.post(PAYPAL_ORDER_API, json=body, headers=headers)
    if response.status_code in (200, 201):
        data = response.json()
        order_id = data["id"]
        approval_url = None
        for link in data["links"]:
            if link["rel"] in ("approve", "payer-action"):
                approval_url = link["href"]
                break
        return order_id, approval_url
    else:
        raise Exception(f"Erreur création ordre PayPal: {response.status_code} {response.text}")

def capture_paypal_order(order_id):
    token = get_paypal_access_token()
    url = f"{PAYPAL_ORDER_API}/{order_id}/capture"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    response = requests.post(url, headers=headers)
    if response.status_code in (200, 201):
        data = response.json()
        if data.get("status") == "COMPLETED":
            return True
        return False
    return False

# Dictionnaire global pour mémoriser temporairement le plan acheté par commande
purchase_orders = {}

@app.route("/purchase_plan/<plan>")
def purchase_plan(plan):
    if plan not in ["1 an", "illimité"]:
        return "Plan non valide", 400
    amount = "10.00" if plan == "1 an" else "40.00"
    try:
        order_id, approval_url = create_paypal_order(amount, "EUR")
        purchase_orders[order_id] = plan
        return redirect(approval_url)
    except Exception as e:
        return f"Erreur: {e}"

@app.route("/paypal_success")
def paypal_success():
    order_id = request.args.get("token", None)
    if not order_id:
        return "Paramètre 'token' manquant dans l'URL."
    success = capture_paypal_order(order_id)
    if success:
        plan = purchase_orders.get(order_id)
        if plan:
            update_activation_after_payment(plan)
            flash(f"Paiement validé pour le plan {plan} !", "success")
        else:
            flash("Paiement validé, mais plan inconnu.", "error")
        return redirect(url_for("saisie"))
    else:
        flash("Paiement non complété.", "error")
        return redirect(url_for("saisie"))

@app.route("/paypal_cancel")
def paypal_cancel():
    flash("Paiement annulé par l'utilisateur.", "error")
    return redirect(url_for("saisie"))

# =============================================================================
# Routes Flask principales
# =============================================================================

@app.before_request
def enforce_activation_flask():
    if request.endpoint not in ("activation", "activate", "purchase_plan", "paypal_success", "paypal_cancel", "change_theme", "trial_expired", "static"):
        if not check_activation():
            return redirect(url_for("activation"))

@app.route("/activation", methods=["GET"])
def activation():
    hw_id = get_hardware_id()
    return render_template_string(
        """
        <!DOCTYPE html>
        <html lang="fr">
        <head>
          <meta charset="UTF-8"/>
          <title>Activation du logiciel</title>
          <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body class="bg-light">
          <div class="container my-5">
            <div class="card shadow">
              <div class="card-body">
                <h2 class="mb-4">Activation du logiciel</h2>
                <p>ID du PC :</p>
                <div class="input-group mb-3">
                  <input type="text" class="form-control" readonly value="{{ hw_id }}">
                </div>
                <p class="text-muted">Pour les plans 1 an et illimité, vous pouvez choisir parmi plusieurs options :</p>
                <form action="{{ url_for('activate') }}" method="POST">
                  <div class="mb-3">
                    <label for="plan" class="form-label">Sélectionnez le Plan :</label>
                    <select name="plan" id="plan" class="form-select">
                      <option value="essai_7jours">Essai Gratuit 7 jours</option>
                      <option value="1 an">1 an (10€)</option>
                      <option value="illimité">Illimité (40€)</option>
                    </select>
                  </div>
                  <div class="mb-3" id="codeDiv">
                    <label for="activation_code" class="form-label">Code d'Activation :</label>
                    <input type="text" name="activation_code" id="activation_code" class="form-control" placeholder="Saisir le code d'activation (si disponible)">
                  </div>
                  <div id="contactOptions" style="display:none;" class="mb-3">
                    <button type="button" class="btn btn-success me-2" onclick="window.location.href='https://api.whatsapp.com/send?phone=212652084735'">WhatsApp</button>
                    <button type="button" class="btn btn-info me-2" onclick="window.location.href='mailto:sastoukadigital@gmail.com'">Email</button>
                    <button type="button" class="btn btn-primary me-2" id="paypalBtn">PayPal</button>
                    <button type="button" class="btn btn-warning" onclick="document.getElementById('activation_code').focus()">Saisir Code</button>
                  </div>
                  <button class="btn btn-primary" type="submit">Valider</button>
                </form>
              </div>
            </div>
          </div>
          <script>
            function updateOptions(){
              var plan = document.getElementById("plan").value;
              var contactDiv = document.getElementById("contactOptions");
              var paypalBtn = document.getElementById("paypalBtn");
              if(plan === "essai_7jours"){
                contactDiv.style.display = "none";
              } else {
                contactDiv.style.display = "block";
                if(plan === "1 an"){
                  paypalBtn.onclick = function(){ window.location.href = "{{ url_for('purchase_plan', plan='1 an') }}"; };
                } else if(plan === "illimité"){
                  paypalBtn.onclick = function(){ window.location.href = "{{ url_for('purchase_plan', plan='illimité') }}"; };
                }
              }
            }
            document.getElementById("plan").addEventListener("change", updateOptions);
            updateOptions();
          </script>
        </body>
        </html>
        """, hw_id=hw_id
    )

@app.route("/activate", methods=["POST"])
def activate():
    plan = request.form.get("plan", "").strip()
    entered_code = request.form.get("activation_code", "").strip()
    hw_id = get_hardware_id()
    if plan == "essai_7jours":
        data = {
            "plan": "essai_7jours",
            "activation_date": datetime.date.today().isoformat()
        }
        with open(ACTIVATION_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f)
        flash("Essai gratuit activé pour 7 jours.", "success")
        return redirect(url_for("saisie"))
    else:
        expected = generate_activation_key_for_user(hw_id, plan)
        if entered_code == expected:
            data = {
                "plan": plan,
                "activation_date": datetime.date.today().isoformat(),
                "activation_code": expected
            }
            with open(ACTIVATION_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f)
            flash(f"Activation validée pour le plan {plan} via saisie de code.", "success")
            return redirect(url_for("saisie"))
        else:
            flash(f"Pour le plan {plan}: Veuillez soit effectuer le paiement via PayPal, saisir un code d'activation valide, ou contacter le développeur par WhatsApp au +212652084735.", "error")
            return redirect(url_for("activation"))

@app.route("/")
def index():
    return redirect(url_for("saisie"))

@app.route("/saisie", methods=["GET", "POST"])
def saisie():
    current_fruit = load_user_theme()
    theme = fruit_themes.get(current_fruit, fruit_themes[DEFAULT_FRUIT])
    if request.method == "POST":
        action = request.form.get("action")
        date_saisie = request.form.get("date_saisie", "").strip()
        agriculteur = request.form.get("agriculteur", "").strip()
        parcelle = request.form.get("parcelle", "").strip()
        produit = request.form.get("produit", "").strip()
        variete = request.form.get("variete", "").strip()
        nb_cueilleurs = request.form.get("nb_cueilleurs", "0").strip()
        nb_indirect = request.form.get("nb_indirect", "0").strip()
        nb_autres = request.form.get("nb_autres", "0").strip()
        nb_caporaux = request.form.get("nb_caporaux", "0").strip()
        poids_total = request.form.get("poids_total", "0").strip()
        ecarts = request.form.get("ecarts", "0").strip()
        try: nb_cueilleurs = int(nb_cueilleurs)
        except: nb_cueilleurs = 0
        try: nb_indirect = int(nb_indirect)
        except: nb_indirect = 0
        try: nb_autres = int(nb_autres)
        except: nb_autres = 0
        total_ouv = nb_cueilleurs + nb_indirect + nb_autres
        try: nb_caporaux = int(nb_caporaux)
        except: nb_caporaux = 0
        try: poids_total = float(poids_total)
        except: poids_total = 0
        try: ecarts = float(ecarts)
        except: ecarts = 0
        poids_global = poids_total + ecarts
        file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "BonLivraison"
            headers = ["Numéro Bon"] + column_order
            for c, h in enumerate(headers, start=1):
                ws.cell(row=1, column=c, value=h)
            wb.create_sheet("HistoriqueRapports")
            wb.save(file_path)
        wb = load_workbook(file_path)
        ws = wb["BonLivraison"]
        next_row = ws.max_row + 1
        num_bon = generate_voucher_number(agriculteur or "AGRI")
        data_dict = {
            "Numéro Bon": num_bon,
            "Date (JJ/MM/AAAA)": date_saisie,
            "Agriculteur": agriculteur,
            "Parcelle": parcelle,
            "Produit": produit,
            "Variété": variete,
            "Nb Ouvriers Cueilleurs": nb_cueilleurs,
            "Nb Ouvriers Indirect": nb_indirect,
            "Nb Ouvriers Autres": nb_autres,
            "Total Ouvriers": total_ouv,
            "Nombre Caporaux": nb_caporaux,
            "Poids Total Cueillis (kg)": poids_total,
            "Écarts (Produit Déchet) en kg": ecarts,
            "Poids Global": poids_global
        }
        row_fields = ["Numéro Bon"] + column_order
        for i, key in enumerate(row_fields, start=1):
            ws.cell(row=next_row, column=i, value=data_dict.get(key, ""))
        wb.save(file_path)
        flash("Enregistré avec succès !", "success")
        if action == "save_pdf":
            pdf_file_name = f"bon_de_livraison_{num_bon}.pdf"
            pdf_path = os.path.join(PDF_LIVRAISON_DIR, pdf_file_name)
            pdf_fields = []
            for col in column_order:
                pdf_val = data_dict.get(col, "")
                pdf_fields.append((col, pdf_val))
            data_for_pdf = {
                "num_bon": f"N° Bon: {num_bon}",
                "fruit": current_fruit,
                "fields": pdf_fields
            }
            try:
                PDFGenerator.generate_delivery_pdf(data_for_pdf, pdf_path)
                flash("PDF généré avec succès !", "success")
            except Exception as e:
                flash(f"Erreur lors de la génération PDF: {e}", "error")
        return redirect(url_for("saisie"))
    now_str = datetime.datetime.now().strftime("%d/%m/%Y")
    return render_template_string(
        """
<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8"/>
  <title>Saisie Bon</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
  <style>
    body { background-color: {{ theme.bg }}; color: {{ theme.fg }}; }
    .navbar-custom { background-color: {{ theme.accent }}; }
    .navbar-custom .navbar-brand, .navbar-custom .nav-link { color: #fff !important; }
    .card { background-color: #ffffffcc; }
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-custom mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="{{ url_for('saisie') }}">Gestion des Récoltes de {{ current_fruit }} ENNAJIHI NAWFAL</a>
    <button class="navbar-toggler text-white" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
      <span class="navbar-toggler-icon">☰</span>
    </button>
    <div class="collapse navbar-collapse" id="navbarNav">
      <ul class="navbar-nav ms-auto">
        <li class="nav-item"><a class="nav-link" href="{{ url_for('bons') }}">Bons antérieurs</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('stats') }}">Statistiques</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('historique') }}">Historique Rapports</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('change_theme') }}">Thème</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='1 an') }}">Acheter 1 an</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='illimité') }}">Acheter Illimité</a></li>
      </ul>
    </div>
  </div>
</nav>
<div class="container">
  <!-- Boutons de contact au niveau de l'interface -->
  <div class="mb-4 text-center">
    <button type="button" class="btn btn-success me-2" onclick="window.location.href='https://api.whatsapp.com/send?phone=212652084735'">WhatsApp</button>
    <button type="button" class="btn btn-info me-2" onclick="window.location.href='mailto:sastoukadigital@gmail.com'">Email</button>
    <button type="button" class="btn btn-primary me-2" onclick="window.location.href='{{ url_for('purchase_plan', plan='1 an') }}'">PayPal (1 an)</button>
    <button type="button" class="btn btn-primary" onclick="window.location.href='{{ url_for('purchase_plan', plan='illimité') }}'">PayPal (Illimité)</button>
  </div>
  <script>
    window.addEventListener('DOMContentLoaded', () => {
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for category, msg in messages %}
            Swal.fire({ icon: "{{ 'error' if category=='error' else 'success' }}", title: "{{ msg|safe }}", timer: 2500 });
          {% endfor %}
        {% endif %}
      {% endwith %}
    });
  </script>
  <div class="card shadow p-4">
    <h3 class="mb-3">Saisie d'un bon de livraison</h3>
    <form method="POST" autocomplete="off">
      <!-- Formulaire inchangé -->
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Date (JJ/MM/AAAA)</label>
        <div class="col-sm-9"><input type="text" name="date_saisie" class="form-control" value="{{ now_str }}"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Agriculteur</label>
        <div class="col-sm-9"><input type="text" name="agriculteur" class="form-control"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Parcelle</label>
        <div class="col-sm-9"><input type="text" name="parcelle" class="form-control"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Produit</label>
        <div class="col-sm-9">
          <select name="produit" class="form-select">
            {% for fruit_name in fruit_themes.keys() %}
              <option value="{{ fruit_name }}">{{ fruit_name }}</option>
            {% endfor %}
          </select>
        </div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Variété</label>
        <div class="col-sm-9"><input type="text" name="variete" class="form-control"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Nb Ouvriers Cueilleurs</label>
        <div class="col-sm-9"><input type="number" name="nb_cueilleurs" class="form-control" value="0"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Nb Ouvriers Indirect</label>
        <div class="col-sm-9"><input type="number" name="nb_indirect" class="form-control" value="0"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Nb Ouvriers Autres</label>
        <div class="col-sm-9"><input type="number" name="nb_autres" class="form-control" value="0"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Nombre Caporaux</label>
        <div class="col-sm-9"><input type="number" name="nb_caporaux" class="form-control" value="0"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Poids Total Cueillis (kg)</label>
        <div class="col-sm-9"><input type="text" name="poids_total" class="form-control" value="0"></div>
      </div>
      <div class="row mb-3">
        <label class="col-sm-3 col-form-label">Écarts (Produit Déchet) en kg</label>
        <div class="col-sm-9"><input type="text" name="ecarts" class="form-control" value="0"></div>
      </div>
      <div class="text-center">
        <button class="btn btn-success me-2" type="submit" name="action" value="save_only">Enregistrer</button>
        <button class="btn btn-primary" type="submit" name="action" value="save_pdf">Enregistrer + PDF</button>
      </div>
    </form>
  </div>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
        """, now_str=now_str, theme=theme, current_fruit=current_fruit, fruit_themes=fruit_themes)

@app.route("/bons")
def bons():
    current_fruit = load_user_theme()
    theme = fruit_themes.get(current_fruit, fruit_themes[DEFAULT_FRUIT])
    search_query = request.args.get("q", "").lower().strip()
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    rows = []
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, sheet_name="BonLivraison")
        for idx, row in df.iterrows():
            row_str = " ".join([str(val).lower() for val in row.values if val is not None])
            if search_query in row_str:
                pt = float(row.get("Poids Total Cueillis (kg)", 0) or 0)
                ec = float(row.get("Écarts (Produit Déchet) en kg", 0) or 0)
                pg = pt + ec
                row_data = {
                    "num_bon": row.get("Numéro Bon", ""),
                    "Date (JJ/MM/AAAA)": row.get("Date (JJ/MM/AAAA)", ""),
                    "Agriculteur": row.get("Agriculteur", ""),
                    "Parcelle": row.get("Parcelle", ""),
                    "Produit": row.get("Produit", ""),
                    "Variété": row.get("Variété", ""),
                    "Nb Ouvriers Cueilleurs": row.get("Nb Ouvriers Cueilleurs", 0),
                    "Nb Ouvriers Indirect": row.get("Nb Ouvriers Indirect", 0),
                    "Nb Ouvriers Autres": row.get("Nb Ouvriers Autres", 0),
                    "Total Ouvriers": row.get("Total Ouvriers", 0),
                    "Nombre Caporaux": row.get("Nombre Caporaux", 0),
                    "Poids Total Cueillis (kg)": pt,
                    "Écarts (Produit Déchet) en kg": ec,
                    "Poids Global": pg,
                    "_idx": idx + 2
                }
                rows.append(row_data)
    return render_template_string("""
<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8"/>
  <title>Bons antérieurs</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
  <style>
    body { background-color: {{ theme.bg }}; color: {{ theme.fg }}; }
    .navbar-custom { background-color: {{ theme.accent }}; }
    .navbar-custom .navbar-brand, .navbar-custom .nav-link { color: #fff !important; }
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-custom mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="{{ url_for('saisie') }}">Gestion des Récoltes - {{ current_fruit }}</a>
    <button class="navbar-toggler text-white" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
      <span class="navbar-toggler-icon">☰</span>
    </button>
    <div class="collapse navbar-collapse" id="navbarNav">
      <ul class="navbar-nav ms-auto">
        <li class="nav-item"><a class="nav-link" href="{{ url_for('saisie') }}">Saisie</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('stats') }}">Statistiques</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('historique') }}">Historique</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('change_theme') }}">Thème</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='1 an') }}">Acheter 1 an</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='illimité') }}">Acheter Illimité</a></li>
      </ul>
    </div>
  </div>
</nav>
<div class="container">
  <script>
    window.addEventListener('DOMContentLoaded', () => {
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for category, msg in messages %}
            Swal.fire({ icon: "{{ 'error' if category=='error' else 'success' }}", title: "{{ msg|safe }}", timer: 2500 });
          {% endfor %}
        {% endif %}
      {% endwith %}
    });
  </script>
  <h3 class="mb-3">Liste des Bons Antérieurs</h3>
  <form method="GET" class="row mb-3">
    <div class="col-auto">
      <input type="text" name="q" class="form-control" placeholder="Recherche Bons..." value="{{ request.args.get('q','') }}">
    </div>
    <div class="col-auto">
      <button class="btn btn-secondary" type="submit">Rechercher</button>
    </div>
  </form>
  <div class="table-responsive">
    <table class="table table-bordered table-striped align-middle">
      <thead>
        <tr>
          <th>N°</th>
          {% for col in column_order %}
            <th>{{ column_abbr.get(col, col) }}</th>
          {% endfor %}
          <th>Actions</th>
        </tr>
      </thead>
      <tbody>
        {% for r in rows %}
        <tr>
          <td>{{ r['num_bon'] }}</td>
          {% for col in column_order %}
            <td>{{ r[col] }}</td>
          {% endfor %}
          <td>
            <a class="btn btn-sm btn-primary" href="{{ url_for('generer_pdf_bon', idx=r['_idx']) }}">PDF</a>
            <a class="btn btn-sm btn-danger" href="{{ url_for('supprimer_bon', idx=r['_idx']) }}" onclick="return confirm('Supprimer ce bon ?');">Supprimer</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
    """, current_fruit=current_fruit, theme=theme, column_order=column_order, column_abbr=column_abbr, rows=rows)

@app.route("/generer_pdf_bon/<int:idx>")
def generer_pdf_bon(idx):
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    if not os.path.exists(file_path):
        flash("Fichier introuvable.", "error")
        return redirect(url_for("bons"))
    df = pd.read_excel(file_path, sheet_name="BonLivraison")
    if idx - 2 < 0 or idx - 2 >= len(df):
        flash("Index invalide.", "error")
        return redirect(url_for("bons"))
    excel_row = df.iloc[idx - 2]
    pt = float(excel_row.get("Poids Total Cueillis (kg)", 0) or 0)
    ec = float(excel_row.get("Écarts (Produit Déchet) en kg", 0) or 0)
    pg = pt + ec
    farmer = str(excel_row.get("Agriculteur", "")).strip()
    unique_num = generate_voucher_number(farmer or "AGRI")
    pdf_file_name = f"bon_de_livraison_{unique_num}.pdf"
    pdf_path = os.path.join(PDF_LIVRAISON_DIR, pdf_file_name)
    data = {"num_bon": "N° Bon: " + unique_num, "fruit": load_user_theme(), "fields": []}
    for col in column_order:
        if col == "Poids Global":
            data["fields"].append((col, pg))
        else:
            data["fields"].append((col, excel_row.get(col, "")))
    PDFGenerator.generate_delivery_pdf(data, pdf_path)
    return send_file(pdf_path, as_attachment=True)

@app.route("/supprimer_bon/<int:idx>")
def supprimer_bon(idx):
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    if not os.path.exists(file_path):
        flash("Fichier Excel introuvable.", "error")
        return redirect(url_for("bons"))
    try:
        wb = load_workbook(file_path)
        ws = wb["BonLivraison"]
        ws.delete_rows(idx)
        wb.save(file_path)
        flash("Bon supprimé avec succès.", "success")
    except Exception as e:
        flash(f"Erreur lors de la suppression : {str(e)}", "error")
    return redirect(url_for("bons"))

@app.route("/stats", methods=["GET", "POST"])
def stats():
    current_fruit = load_user_theme()
    theme = fruit_themes.get(current_fruit, fruit_themes[DEFAULT_FRUIT])
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    df = pd.DataFrame()
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, sheet_name="BonLivraison")
        if "Date (JJ/MM/AAAA)" in df.columns:
            df["Date"] = pd.to_datetime(df["Date (JJ/MM/AAAA)"], format="%d/%m/%Y", errors="coerce")
        else:
            df["Date"] = None
        df["Mois"] = df["Date"].dt.month
        df["Annee"] = df["Date"].dt.year
        for c in ["Poids Total Cueillis (kg)", "Écarts (Produit Déchet) en kg", 
                  "Nb Ouvriers Cueilleurs", "Nb Ouvriers Indirect", "Nb Ouvriers Autres", "Nombre Caporaux"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
            else:
                df[c] = 0
        df["Poids Global"] = df["Poids Total Cueillis (kg)"] + df["Écarts (Produit Déchet) en kg"]
        df["Total Ouvriers"] = df["Nb Ouvriers Cueilleurs"] + df["Nb Ouvriers Indirect"] + df["Nb Ouvriers Autres"]
    start_date_str = request.form.get("start_date", "")
    end_date_str = request.form.get("end_date", "")
    graph_column = request.form.get("graph_column", "Poids Total Cueillis (kg)")
    x_axis = request.form.get("x_axis", "Date (JJ/MM/AAAA)")
    selected_checkboxes = request.form.getlist("checkbox_fields")
    if start_date_str:
        try:
            sd = datetime.datetime.strptime(start_date_str, "%d/%m/%Y")
            df = df[df["Date"] >= sd]
        except:
            pass
    if end_date_str:
        try:
            ed = datetime.datetime.strptime(end_date_str, "%d/%m/%Y")
            df = df[df["Date"] <= ed]
        except:
            pass
    if not df.empty and x_axis in df.columns and graph_column in df.columns:
        group_data = df.groupby(x_axis, as_index=False).agg({graph_column: "sum"})
    else:
        group_data = pd.DataFrame({x_axis:[], graph_column:[]})
    
    # Graphiques principaux dynamiques (Histogramme et Donut) pour l'ensemble
    bar_img = None
    pie_img = None
    if not group_data.empty and x_axis in group_data.columns and graph_column in group_data.columns:
        fig_bar, ax_bar = plt.subplots(figsize=(5,4), dpi=100)
        x_values = group_data[x_axis].astype(str).tolist()
        y_values = group_data[graph_column].tolist()
        accent_color = theme["accent"]
        ax_bar.bar(x_values, y_values, color=accent_color)
        ax_bar.set_title(f"Histogramme selon {x_axis}")
        ax_bar.set_xlabel(x_axis)
        ax_bar.set_ylabel(graph_column)
        plt.setp(ax_bar.get_xticklabels(), rotation=45)
        fig_bar.tight_layout()
        bar_buffer = BytesIO()
        fig_bar.savefig(bar_buffer, format='png')
        plt.close(fig_bar)
        bar_buffer.seek(0)
        bar_img = base64.b64encode(bar_buffer.getvalue()).decode('utf-8')
        
        fig_pie, ax_pie = plt.subplots(figsize=(5,4), dpi=100)
        if sum(y_values) > 0:
            wedges, texts, autotexts = ax_pie.pie(y_values, autopct='%1.1f%%', startangle=90, wedgeprops={'width':0.3})
            ax_pie.set_title(f"Camembert selon {x_axis}")
            ax_pie.legend(wedges, x_values, title=x_axis, loc="center left", bbox_to_anchor=(1, 0.5), fontsize=10, title_fontsize=12)
            fig_pie.tight_layout()
            pie_buffer = BytesIO()
            fig_pie.savefig(pie_buffer, format='png')
            plt.close(fig_pie)
            pie_buffer.seek(0)
            pie_img = base64.b64encode(pie_buffer.getvalue()).decode('utf-8')
        else:
            plt.close(fig_pie)
    
    # Pour chaque case cochée, créer un histogramme et un donut côte à côte dans une figure dynamique
    selected_stats = {}
    for field in selected_checkboxes:
        try:
            # Agrégation sur le champ sélectionné (somme du graph_column)
            group = df.groupby(field)[graph_column].sum().reset_index()
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10,4), dpi=100)
            # Histogramme
            ax1.bar(group[field].astype(str), group[graph_column], color=theme["accent"])
            ax1.set_title(f"Histogramme par {field}")
            ax1.set_xlabel(field)
            ax1.set_ylabel(graph_column)
            plt.setp(ax1.get_xticklabels(), rotation=45)
            # Donut chart
            total = group[graph_column].sum()
            if total > 0:
                wedges, texts, autotexts = ax2.pie(group[graph_column], autopct='%1.1f%%', startangle=90, wedgeprops={'width':0.3})
                ax2.set_title(f"Donut par {field}")
                ax2.legend(wedges, group[field].astype(str), title=field, loc="center left", bbox_to_anchor=(1, 0.5), fontsize=10, title_fontsize=12)
            else:
                ax2.text(0.5, 0.5, "Aucune donnée", ha="center", va="center")
            fig.tight_layout()
            buf = BytesIO()
            fig.savefig(buf, format='png')
            plt.close(fig)
            buf.seek(0)
            chart_img = base64.b64encode(buf.getvalue()).decode('utf-8')
            selected_stats[field] = {"table": group.to_dict(orient="records"), "chart": chart_img}
        except Exception as e:
            selected_stats[field] = {"table": [], "chart": ""}
    
    if request.method == "POST":
        action = request.form.get("action")
        if action == "generate_pdf_stats":
            report_num = generate_report_number("Rapport")
            pdf_file_name = f"Rapport_{graph_column}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
            pdf_path = os.path.join(PDF_STATS_DIR, pdf_file_name)
            from reportlab.platypus import Table, TableStyle, Spacer, Paragraph, PageBreak
            styles = getSampleStyleSheet()
            styles["Normal"].fontSize = 14
            styles["Normal"].alignment = TA_JUSTIFY
            styles["Title"].fontSize = 14
            styles["Title"].alignment = TA_CENTER
            styles["Heading2"].fontSize = 14
            styles["Heading2"].alignment = TA_CENTER
            elements = []
            # Tableau et graphiques principaux
            elements.append(Paragraph(f"Rapport de statistiques du {start_date_str} au {end_date_str}", styles["Heading2"]))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph("N° Rapport: " + report_num, styles["Heading2"]))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Statistiques selon '{x_axis}'", styles['Heading2']))
            available_width = landscape(A4)[0] - 40
            col_widths = [available_width * 0.5, available_width * 0.5]
            data_table = [[x_axis, graph_column]]
            for _, row in group_data.iterrows():
                data_table.append([str(row[x_axis]), row[graph_column]])
            table_pdf = Table(data_table, repeatRows=1, colWidths=col_widths)
            table_pdf.hAlign = 'CENTER'
            table_pdf.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.darkblue),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 14),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.beige]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
            ]))
            elements.append(table_pdf)
            elements.append(Spacer(1, 12))
            # Insertion des graphiques principaux
            from reportlab.platypus import Image
            images = []
            if bar_img:
                bar_buffer = BytesIO(base64.b64decode(bar_img))
                images.append(Image(bar_buffer, width=available_width*0.48, height=250))
            if pie_img:
                pie_buffer = BytesIO(base64.b64decode(pie_img))
                images.append(Image(pie_buffer, width=available_width*0.48, height=250))
            if images:
                table_images = Table([images], colWidths=[available_width*0.5, available_width*0.5])
                table_images.hAlign = 'CENTER'
                elements.append(table_images)
            elements.append(PageBreak())
            # Pour chaque case cochée, insertion d'un tableau et d'un graphique avec saut de page
            for field, stats_data in selected_stats.items():
                elements.append(Paragraph(f"Statistiques pour {field} :", styles["Heading2"]))
                table_data = [[field, graph_column]]
                for row in stats_data["table"]:
                    table_data.append([str(row[field]), row[graph_column]])
                stat_table = Table(table_data, repeatRows=1)
                stat_table.hAlign = 'CENTER'
                stat_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
                ]))
                elements.append(stat_table)
                elements.append(Spacer(1, 12))
                if stats_data["chart"]:
                    buf = BytesIO(base64.b64decode(stats_data["chart"]))
                    elements.append(Image(buf, width=available_width, height=250))
                elements.append(PageBreak())
            PDFGenerator.generate_stats_pdf(elements, pdf_path)
            wb = load_workbook(file_path)
            if "HistoriqueRapports" not in wb.sheetnames:
                ws_hist = wb.create_sheet("HistoriqueRapports")
                ws_hist.append(["N°", "Type", "Date", "Chemin"])
            else:
                ws_hist = wb["HistoriqueRapports"]
            max_row = ws_hist.max_row
            idx_rapport = max_row
            ws_hist.append([idx_rapport, "Statistiques", datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"), pdf_path])
            wb.save(file_path)
            return send_file(pdf_path, as_attachment=True)
    return render_template_string("""
<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8"/>
  <title>Statistiques</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
  <style>
    body { background-color: {{ theme.bg }}; color: {{ theme.fg }}; }
    .navbar-custom { background-color: {{ theme.accent }}; }
    .navbar-custom .navbar-brand, .navbar-custom .nav-link { color: #fff !important; }
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-custom mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="{{ url_for('saisie') }}">Gestion des Récoltes - {{ current_fruit }}</a>
    <button class="navbar-toggler text-white" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
      <span class="navbar-toggler-icon">☰</span>
    </button>
    <div class="collapse navbar-collapse" id="navbarNav">
      <ul class="navbar-nav ms-auto">
        <li class="nav-item"><a class="nav-link" href="{{ url_for('saisie') }}">Saisie</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('bons') }}">Bons antérieurs</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('historique') }}">Historique</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('change_theme') }}">Thème</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='1 an') }}">Acheter 1 an</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='illimité') }}">Acheter Illimité</a></li>
      </ul>
    </div>
  </div>
</nav>
<div class="container">
  <!-- Boutons de contact dans la rubrique Statistiques -->
  <div class="mb-4 text-center">
    <button type="button" class="btn btn-success me-2" onclick="window.location.href='https://api.whatsapp.com/send?phone=212652084735'">WhatsApp</button>
    <button type="button" class="btn btn-info me-2" onclick="window.location.href='mailto:sastoukadigital@gmail.com'">Email</button>
    <button type="button" class="btn btn-primary me-2" onclick="window.location.href='{{ url_for('purchase_plan', plan='1 an') }}'">PayPal (1 an)</button>
    <button type="button" class="btn btn-primary" onclick="window.location.href='{{ url_for('purchase_plan', plan='illimité') }}'">PayPal (Illimité)</button>
  </div>
  <div class="card shadow p-4 mb-4">
    <h3>Filtre Statistiques</h3>
    <form method="POST">
      <div class="row mb-3">
        <div class="col">
          <label>Date début (JJ/MM/AAAA)</label>
          <input type="text" name="start_date" class="form-control" value="{{ request.form.get('start_date','')}}">
        </div>
        <div class="col">
          <label>Date fin (JJ/MM/AAAA)</label>
          <input type="text" name="end_date" class="form-control" value="{{ request.form.get('end_date','')}}">
        </div>
      </div>
      <div class="row mb-3">
        <div class="col">
          <label>Y-axis</label>
          <select name="graph_column" class="form-select">
            {% for col in all_columns_extended %}
              <option value="{{ col }}" {% if col==graph_column %}selected{% endif %}>{{ col }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col">
          <label>X-axis</label>
          <select name="x_axis" class="form-select">
            {% for col in all_columns_extended %}
              <option value="{{ col }}" {% if col==x_axis %}selected{% endif %}>{{ col }}</option>
            {% endfor %}
          </select>
        </div>
      </div>
      <div class="mb-3">
        <label>Champs supplémentaires :</label><br>
        {% for field in ["Mois", "Annee", "Agriculteur", "Produit", "Variété"] %}
          <div class="form-check form-check-inline">
            <input class="form-check-input" type="checkbox" name="checkbox_fields" value="{{ field }}" id="{{ field }}"
                   {% if field in selected_checkboxes %}checked{% endif %}>
            <label class="form-check-label" for="{{ field }}">{{ field }}</label>
          </div>
        {% endfor %}
      </div>
      <button class="btn btn-info" type="submit">Mettre à jour</button>
      <button class="btn btn-warning" type="submit" name="action" value="generate_pdf_stats">Générer PDF</button>
    </form>
  </div>
  {% if bar_img %}
  <div class="row mb-4">
    <div class="col-md-6">
      <img src="data:image/png;base64,{{ bar_img }}" class="img-fluid" alt="Histogramme">
    </div>
    <div class="col-md-6">
      {% if pie_img %}
      <img src="data:image/png;base64,{{ pie_img }}" class="img-fluid" alt="Camembert">
      {% endif %}
    </div>
  </div>
  {% endif %}
  <div class="card shadow p-4 mb-4">
    <h3>Tableau récapitulatif</h3>
    <table class="table table-bordered table-striped mt-3">
      <thead>
        <tr>
          <th>{{ x_axis }}</th>
          <th>{{ graph_column }}</th>
        </tr>
      </thead>
      <tbody>
        {% for _, row in group_data.iterrows() %}
        <tr>
          <td>{{ row[x_axis] }}</td>
          <td>{{ row[graph_column] }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% for field, stat in selected_stats.items() %}
  <div class="card shadow p-4 mb-4">
    <h3>Statistiques pour {{ field }}</h3>
    <table class="table table-bordered table-striped mt-3">
      <thead>
        <tr>
          <th>{{ field }}</th>
          <th>{{ graph_column }}</th>
        </tr>
      </thead>
      <tbody>
        {% for row in stat.table %}
        <tr>
          <td>{{ row[field] }}</td>
          <td>{{ row[graph_column] }}</td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    {% if stat.chart %}
    <div class="mt-3 text-center">
      <img src="data:image/png;base64,{{ stat.chart }}" class="img-fluid" alt="Graphique pour {{ field }}">
    </div>
    {% endif %}
  </div>
  {% endfor %}
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
    """, current_fruit=current_fruit, theme=theme, all_columns_extended=all_columns_extended, graph_column=graph_column, x_axis=x_axis, selected_checkboxes=selected_checkboxes, group_data=group_data, bar_img=bar_img, pie_img=pie_img, selected_stats=selected_stats)

@app.route("/historique")
def historique():
    current_fruit = load_user_theme()
    theme = fruit_themes.get(current_fruit, fruit_themes[DEFAULT_FRUIT])
    file_path = os.path.join(EXCEL_DIR, "enregistrements.xlsx")
    report_history = []
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if "HistoriqueRapports" in wb.sheetnames:
            ws = wb["HistoriqueRapports"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    entry = {"idx": row[0], "type": row[1], "date": row[2], "pdf_path": row[3]}
                    report_history.append(entry)
    search_query = request.args.get("q", "").lower().strip()
    if search_query:
        filtered = []
        for rh in report_history:
            concat = (str(rh["idx"]) + rh["type"] + str(rh["date"]) + str(rh["pdf_path"])).lower()
            if search_query in concat:
                filtered.append(rh)
        report_history = filtered
    return render_template_string("""
<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8"/>
  <title>Historique Rapports</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/sweetalert2@11"></script>
  <style>
    body { background-color: {{ theme.bg }}; color: {{ theme.fg }}; }
    .navbar-custom { background-color: {{ theme.accent }}; }
    .navbar-custom .navbar-brand, .navbar-custom .nav-link { color: #fff !important; }
  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-custom mb-4">
  <div class="container-fluid">
    <a class="navbar-brand" href="{{ url_for('saisie') }}">Gestion des Récoltes - {{ current_fruit }}</a>
    <button class="navbar-toggler text-white" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
      <span class="navbar-toggler-icon">☰</span>
    </button>
    <div class="collapse navbar-collapse" id="navbarNav">
      <ul class="navbar-nav ms-auto">
        <li class="nav-item"><a class="nav-link" href="{{ url_for('saisie') }}">Saisie</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('bons') }}">Bons antérieurs</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('stats') }}">Statistiques</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('change_theme') }}">Thème</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='1 an') }}">Acheter 1 an</a></li>
        <li class="nav-item"><a class="nav-link" href="{{ url_for('purchase_plan', plan='illimité') }}">Acheter Illimité</a></li>
      </ul>
    </div>
  </div>
</nav>
<div class="container">
  <script>
    window.addEventListener('DOMContentLoaded', () => {
      {% with messages = get_flashed_messages(with_categories=true) %}
        {% if messages %}
          {% for category, msg in messages %}
            Swal.fire({ icon: "{{ 'error' if category=='error' else 'success' }}", title: "{{ msg|safe }}", timer: 2500 });
          {% endfor %}
        {% endif %}
      {% endwith %}
    });
  </script>
  <h3 class="mb-3">Historique des Rapports</h3>
  <form method="GET" class="row mb-3">
    <div class="col-auto">
      <input type="text" name="q" class="form-control" placeholder="Rechercher..." value="{{ request.args.get('q','') }}">
    </div>
    <div class="col-auto">
      <button class="btn btn-secondary">Rechercher</button>
    </div>
  </form>
  <table class="table table-bordered table-striped">
    <thead>
      <tr>
        <th>N°</th>
        <th>Type</th>
        <th>Date</th>
        <th>Chemin</th>
        <th>Action</th>
      </tr>
    </thead>
    <tbody>
      {% for r in report_history %}
      <tr>
        <td>{{ r["idx"] }}</td>
        <td>{{ r["type"] }}</td>
        <td>{{ r["date"] }}</td>
        <td>{{ r["pdf_path"] }}</td>
        <td>
          {% if r["pdf_path"] and r["pdf_path"]|length > 0 %}
            <a class="btn btn-sm btn-primary" href="{{ url_for('afficher_rapport', pdfpath=r['pdf_path']) }}">Afficher</a>
          {% else %}
            -
          {% endif %}
        </td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>
    """, current_fruit=current_fruit, theme=theme, report_history=report_history)

@app.route("/afficher_rapport")
def afficher_rapport():
    pdf_path = request.args.get("pdfpath", "")
    if not os.path.exists(pdf_path):
        flash("Fichier PDF introuvable.", "error")
        return redirect(url_for("historique"))
    return send_file(pdf_path, as_attachment=False)

@app.route("/change_theme", methods=["GET", "POST"])
def change_theme():
    if request.method == "POST":
        chosen_fruit = request.form.get("fruit", DEFAULT_FRUIT)
        save_user_theme(chosen_fruit)
        flash(f"Thème '{chosen_fruit}' enregistré.", "success")
        return redirect(url_for("saisie"))
    current_fruit = load_user_theme()
    return render_template_string("""
<!DOCTYPE html>
<html lang="fr">
<head>
  <meta charset="UTF-8">
  <title>Choisir un fruit (Thème)</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
<div class="container my-5">
  <h3>Choisir un fruit (Thème)</h3>
  <form method="POST" class="mb-3">
    <select name="fruit" class="form-select">
      {% for fr in fruit_themes.keys() %}
        <option value="{{ fr }}" {% if fr==current_fruit %}selected{% endif %}>{{ fr }}</option>
      {% endfor %}
    </select>
    <button class="btn btn-primary mt-3" type="submit">Enregistrer</button>
  </form>
  <a class="btn btn-secondary" href="{{ url_for('saisie') }}">Retour</a>
</div>
</body>
</html>
    """, fruit_themes=fruit_themes, current_fruit=current_fruit)

# =============================================================================
# Lancement du serveur Flask
# =============================================================================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
